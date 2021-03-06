import os
import re
from datetime import date, datetime

from bottle import route, run, template
from bottle import post, request, Bottle
from bottle import static_file
from openpyxl import Workbook
from openpyxl.compat import range
from openpyxl.utils import get_column_letter
from pony.orm import *

from database.management import DBM
from database.models import db, Complaint

dbm = DBM()
app = Bottle()

@app.route('/')
def displayForm():
    return template('templates/form.tpl')

@app.route('/delete')
def deleteAllResolve():
    dbm.deleteAllResolve()
    return template('templates/msg.tpl', filePath="", message="All Resolved cases have been deleted")

@app.post('/query/complaint')
def getResolve():
    info = dbm.getResolve(request.forms.get('complaint_number'))
    r = []
    if len(info.keys()) != 0:
        r.append(info)

    return template('templates/resolve.tpl', infoList=r)

@app.post('/query/inspection')
def getResolveByLastInspection():
    startDate = datetime.strptime(request.forms.get('startDate'), "%Y-%m-%d")
    endDate = datetime.strptime(request.forms.get('endDate'), "%Y-%m-%d")
    return template('templates/resolve.tpl', infoList = \
        dbm.getResolveByInspectionDateRange(startDate, endDate))

@app.post('/query/disposition')
def getResolveByDisposition():
    startDate = datetime.strptime(request.forms.get('startDate'), "%Y-%m-%d")
    endDate = datetime.strptime(request.forms.get('endDate'), "%Y-%m-%d")
    return template('templates/resolve.tpl', infoList = \
        dbm.getResolveByDisposition(startDate, endDate))

@app.post('/query/category')
def getResolveByCategory():
    return template('templates/resolve.tpl', infoList = \
                        dbm.getResolveByCategoryCode(request.forms.get('category')))

@app.route('/output/<filename:path>')
def downloadLink(filename):
    return static_file(filename, root='output/', download=filename)

@app.route('/clearup')
def clearCache():
    folder = 'output/'
    for entry in os.listdir(folder):
        filepath = os.path.join(folder, entry)
        try:
            if os.path.isfile(filepath):
                os.unlink(filepath)
        except e:
            print e
    return template('templates/msg.tpl', filePath="", message="Successfully Cleared")

@app.route('/export')
def export():
    fileName = date.today().strftime("%m-%d-%Y") + "_Resolved.xlsx"        
    success = exportResolve(fileName)
    if success:
        return template('templates/msg.tpl', filePath="output/" + fileName, message="Successfully Exported")
    else:
        return template('templates/msg.tpl', filePath="", message="There is no resolved cases in database")

@app.route('/static/script/<filecategory:path>/<filename:path>')
def sendStaticFile(filecategory, filename):
    return static_file(filecategory + "/" + filename, root='static/script/')

#@app.route('/static/script/css/<cssFile:path>')
#def sendCSSFile(cssFile):
#    return static_file(cssFile, root='static/script/css/')

def exportResolve(fileName):
    if not os.path.exists("output"):
        os.makedirs("output")

    disPattern = re.compile("(\xa0){0,}(?P<dis>[0-9/]+).*")
    savePath = "output/" + fileName
    b = Workbook()
    s = b.active
    resolve = dbm.getAllResolve()
    for case in resolve:
        case['Disposition Date'] = disPattern.match(case['Disposition']).group(2)
    dataCounts = len(resolve)
    if dataCounts == 0:
        return False   
    keys = sorted(resolve[0].keys())
    keySize = len(keys)

    # Write column head
    for col_index in range(1, keySize + 1):
        s["%s1"%get_column_letter(col_index)] = keys[col_index - 1]

    # Write data from second row
    i = 0
    for row_index in range(2, dataCounts + 2):
        # Reduce index calculation when access a data
        data = resolve[i]
        for col_index in range(1, keySize + 1):
            s["%s%s"%(get_column_letter(col_index), row_index)] = \
                        data[keys[col_index - 1]]
                
        i += 1

    b.save(filename = savePath)
    return True

app.run(host='localhost', port=8080, debug=True)
